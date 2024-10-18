import pandas as pd
from sqlalchemy import create_engine
import os
from datetime import timedelta, datetime
import locale
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
import win32com.client as win32
import numpy as np
import time
import calendar

locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
data_atual = datetime.now().strftime("%d-%m-%Y")

def consulta_sql_dados(conexao_sql, consulta_sql):
    try:
        engine = create_engine(conexao_sql)
        df = pd.read_sql(consulta_sql, engine)
        return df
    except Exception as e:
        print(f'Ocorreu um erro ao executar a consulta SQL: {e}')
        return None

def formatar_excel(arquivo_excel, df):
    wb = load_workbook(arquivo_excel)
    ws = wb.active

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    
    header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") 
    header_font = Font(color="000000", bold=True, size=12)

    for cell in ws[4]:  
        cell.fill = header_fill
        cell.font = header_font
        
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.font = Font(size=8) 

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)       
        ws.column_dimensions[column].width = adjusted_width

    wb.save(arquivo_excel)

def salvar_pdfs_por_executivos(df, coluna_executivos, pasta_destino):
    executivos = df[coluna_executivos].unique()
    os.makedirs(pasta_destino, exist_ok=True)

    for executivo in executivos:
        try:
            day_py = datetime.now().day

            if day_py == 1:
                mes_passado_py = (datetime.now() - relativedelta(months=2)).strftime('%B/%y')
                mes_atual_py = (datetime.now() - relativedelta(months=1)).strftime('%B/%y')
            else:
                mes_passado_py = (datetime.now() - relativedelta(months=1)).strftime('%B/%y')
                mes_atual_py = datetime.now().strftime('%B/%y')

            df_executivo = df[df[coluna_executivos] == executivo]
            nome_arquivo_excel = os.path.join(pasta_destino, f'{executivo} - DETALHE - {data_atual}.xlsx')
            colunas_existentes = [col for col in ['CD_ENTIDADE', 'DS_FANTASIA', mes_passado_py, mes_atual_py, 'Projeção', 'STATUS_VENDA', 'EXECUTIVO'] if col in df_executivo.columns]
            df_executivo = df_executivo[colunas_existentes]
            df_executivo.to_excel(nome_arquivo_excel, index=False)
            formatar_excel(nome_arquivo_excel, df_executivo) 
            salvar_como_pdf(nome_arquivo_excel, executivo)

            if os.path.exists(nome_arquivo_excel):
                os.remove(nome_arquivo_excel)
            print(f'Arquivo PDF do executivo {executivo} salvo com sucesso!')
        except Exception as e:
            print(f'Ocorreu um erro ao salvar o arquivo PDF do executivo {executivo}: {e}')

def salvar_como_pdf(caminho_excel, nome_executivo):
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(caminho_excel)
        ws = wb.Sheets(1)
        ws.PageSetup.Zoom = False 
        ws.PageSetup.FitToPagesWide = 1  
        ws.PageSetup.FitToPagesTall = False 
        caminho_pdf = caminho_excel.replace(".xlsx", ".pdf")
        wb.ExportAsFixedFormat(0, caminho_pdf)
        wb.Close(False)
        excel.Quit()
        print(f'Arquivo PDF do executivo {nome_executivo} salvo com sucesso!')
    except Exception as e:
        print(f'Ocorreu um erro ao salvar o arquivo PDF do executivo {nome_executivo}: {e}')
    finally:
        excel.Quit()

def contar_dias_uteis_mes(ano, mes, dia_limite=None):
    if dia_limite is None:
        _, ultimo_dia_mes = calendar.monthrange(ano, mes)
        dia_limite = ultimo_dia_mes
    
    dias_uteis = 0
    for dia in range(1, dia_limite + 1):
        dia_semana = calendar.weekday(ano, mes, dia)
        if dia_semana < 5:
            dias_uteis += 1
            
    return dias_uteis

def obter_mes_passado():
    ano_atual, mes_atual = time.localtime().tm_year, time.localtime().tm_mon
    if mes_atual == 1:
        return ano_atual - 1, 12
    else:
        return ano_atual, mes_atual - 1

def contar_dias_uteis():
    ano_atual, mes_atual, dia_atual = time.localtime().tm_year, time.localtime().tm_mon, time.localtime().tm_mday
    ano_passado, mes_passado = obter_mes_passado()
    dias_uteis_mes_passado = contar_dias_uteis_mes(ano_passado, mes_passado)
    dias_uteis_mes_atual = contar_dias_uteis_mes(ano_atual, mes_atual, dia_limite=dia_atual)

    return dias_uteis_mes_passado, dias_uteis_mes_atual

def etapas_power_query(df):
    mes_passado_py, mes_atual_py = calcular_meses()
    df['DATA'] = pd.to_datetime(df['DATA'], errors='coerce')
    df['DATA'] = df['DATA'].dt.strftime('%B/%y')    
    df_pivot = df.pivot_table(index=['CD_ENTIDADE', 'DS_FANTASIA', 'STATUS_VENDA', 'EXECUTIVO'],
                              columns='DATA', values='VOLUME', aggfunc='sum').reset_index()
    
    colunas_ordenadas = ['CD_ENTIDADE', 'DS_FANTASIA', mes_passado_py, mes_atual_py, 'STATUS_VENDA', 'EXECUTIVO']
    df_pivot = df_pivot.reindex(columns=[col for col in colunas_ordenadas if col in df_pivot.columns], fill_value=0)
    executivos_excluir = [
        'FORÇA DE VENDAS INTERNA', 'GPS CORRETAGENS E ADMINISTRACAO DE SEGUROS', 'ISABELA FERNANDA DA SILVA AMBROZIO',
        'JULIANA COSTA MARTINS', 'JULIANA DE OLIVEIRA RODRIGUES', 'MARIA IZABEL OLIVEIRA TRINDADE', 'NOSSA REDE',
        'THAMIRIS FIGUEIREDO DOS SANTOS AVELINO', 'THIAGO FERREIRA MELHEIRO'
    ]
    df_filtrado = df_pivot[~df_pivot['EXECUTIVO'].isin(executivos_excluir)]    
    df_classificado = df_filtrado.sort_values(by=mes_passado_py, ascending=False)

    return df_classificado

def adicionar_projecao(df):
    mes_passado_py, mes_atual_py = calcular_meses()

    if mes_passado_py in df.columns and mes_atual_py in df.columns:
        dias_passados = datetime.now().day        
        ano_atual, mes_atual = datetime.now().year, datetime.now().month
        dias_no_mes = calendar.monthrange(ano_atual, mes_atual)[1]
        df[mes_atual_py] = df[mes_atual_py].fillna(0) 
        df['Media_Diaria_Atual'] = df[mes_atual_py] / dias_passados
        df['Projeção'] = round(df['Media_Diaria_Atual'] * dias_no_mes)
        df = df.drop(columns=['Media_Diaria_Atual'])
    else:
        print(f"As colunas {mes_passado_py} e {mes_atual_py} não foram encontradas no DataFrame.")   

    return df

def calcular_meses():
    dia_atual = datetime.now().day    
    if dia_atual == 1:
        mes_passado_py = (datetime.now() - relativedelta(months=2)).strftime('%B/%y')
        mes_atual_py = (datetime.now() - relativedelta(months=1)).strftime('%B/%y')
    else:
        mes_passado_py = (datetime.now() - relativedelta(months=1)).strftime('%B/%y')
        mes_atual_py = datetime.now().strftime('%B/%y')
    return mes_passado_py, mes_atual_py


usuario = 'joao.ruiz'
senha = 'pSLOCgBgy5tB'
servidor = 'sodreproducao.public.608eda075a62.database.windows.net,3342'
banco_de_dados = 'toxicologico'
conexao_sql = f'mssql+pyodbc://{usuario}:{senha}@{servidor}/{banco_de_dados}?driver=ODBC+Driver+17+for+SQL+Server'

consulta_sql = """
DECLARE @DiaAtual INT = DAY(GETDATE());

DECLARE @MesPassado DATE;
DECLARE @MesAtual DATE;

IF @DiaAtual = 1
BEGIN    
    SET @MesPassado = CONVERT(DATE, CONCAT(MONTH(DATEADD(MONTH, -2, GETDATE())), '/', 01, '/', YEAR(GETDATE())));
    SET @MesAtual = EOMONTH(DATEADD(MONTH, -1, GETDATE()))
END
ELSE
BEGIN
    SET @MesPassado = CONVERT(DATE, CONCAT(MONTH(DATEADD(MONTH, -1, GETDATE())), '/', 01, '/', YEAR(GETDATE())));
    SET @MesAtual = EOMONTH(GETDATE())
END

SELECT 
    CASE 
        WHEN v.cd_empresa IS NULL 
        THEN v.cd_laboratorio 
        ELSE v.cd_empresa 
    END AS CD_ENTIDADE,
    e.DS_FANTASIA, 
    CAST(CONCAT(MONTH(v.dt_venda), '/', 01, '/', YEAR(GETDATE())) AS DATE) AS DATA,
    COUNT(v.ds_etiqueta) AS VOLUME, 
    CASE 
        WHEN DATEDIFF(DAY, v.dt_venda, GETDATE()) <= 90
            AND e.dt_cadastro >= '2022-12-31' 
            AND v.ds_finalidade <> 'TERCEIRO'
        THEN 'VENDA NOVA'
        ELSE 'VENDA MANUTENÇÃO'
    END AS STATUS_VENDA,
    f.ds_fantasia AS EXECUTIVO
FROM powerbi.consultavendas v 
LEFT JOIN morales.dbo.tbl_entidades e 
    ON (CASE 
            WHEN v.cd_empresa IS NULL 
            THEN v.cd_laboratorio 
            ELSE v.cd_empresa
        END) = e.cd_entidade
LEFT JOIN powerbi.executivos f 
    ON e.cd_vendedor = f.cd_vendedor 
WHERE v.dt_venda BETWEEN @MesPassado AND @MesAtual
GROUP BY 
    CASE 
        WHEN v.cd_empresa IS NULL 
        THEN v.cd_laboratorio 
        ELSE v.cd_empresa 
    END, 
    e.DS_FANTASIA,
    CAST(CONCAT(MONTH(v.dt_venda), '/', 01, '/', YEAR(GETDATE())) AS DATE),
    f.ds_fantasia,
    CASE 
        WHEN DATEDIFF(DAY, v.dt_venda, GETDATE()) <= 90 
            AND e.dt_cadastro >= '2022-12-31'
            AND v.ds_finalidade <> 'TERCEIRO'
        THEN 'VENDA NOVA'
        ELSE 'VENDA MANUTENÇÃO'
    END
ORDER BY 
    COUNT(v.ds_etiqueta) DESC
"""

destino_arquivo = r'C:\Users\wkasouto\OneDrive\OneDrive - Laboratorio Morales LTDA\Felipe Corradi\Relatórios Executivos\Relatório Segunda-Feira'
df = consulta_sql_dados(conexao_sql, consulta_sql)
df_transformado = etapas_power_query(df)
df_com_projecao = adicionar_projecao(df_transformado)

contar_dias_uteis()

salvar_pdfs_por_executivos(df_com_projecao, 'EXECUTIVO', destino_arquivo)
